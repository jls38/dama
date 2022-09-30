import logging
import re
from dataclasses import dataclass
from typing import Optional

from openpyxl import utils as opxl_utils
from openpyxl.utils import cell as opxl_cell
from openpyxl.worksheet.worksheet import Worksheet as OpxlWorksheet

from DamaLib.common.decorators.check import check_method_input, check_dataclass_input
from constants.constant import MAX_COLUMN_EXL, MAX_ROW_EXL

log = logging.getLogger(__name__)

@check_dataclass_input
@dataclass
class DMSelect:
    worksheet:OpxlWorksheet
    start_cell:str
    end_cell:Optional[str] = None

    def __post_init__(self):
        pass

@check_dataclass_input
@dataclass
class _CoordinateParam:
    coordinate:str
    col_letter = str()
    col_number = int()
    row_number = int()

    def __post_init__(self):
        if not CellsCoordinates._isValidExcelCell(self.coordinate):
            raise ValueError("cell coordinate not valid")

        self.col_letter = opxl_cell.coordinate_from_string(self.coordinate)[0]
        self.col_number = int(opxl_cell.column_index_from_string(self.col_letter))
        self.row_number = int(opxl_cell.coordinate_from_string(self.coordinate)[1])

@check_dataclass_input
@dataclass
class CellsSelection:
    worksheet:OpxlWorksheet
    start_cell:_CoordinateParam
    end_cell:_CoordinateParam
    rows_range:range 
    columns_range:range

    def __post_init__(self):
        pass

class CellsCoordinates(object):
    def __init__() -> None:
        pass

    @check_method_input('')
    @staticmethod
    def set_up_selection(select:DMSelect) -> CellsSelection:
        """
        if not type(select) == DMSelect:
            print('select: ', select)
            raise TypeError('select must be DMSelect type')
        """
        selection = CellsSelection

        # Set worksheet
        if not CellsCoordinates._isValidExcelWorksheet(select.worksheet):
            raise Exception('Worksheet not valid')
        selection.worksheet = select.worksheet

        # Set start_cell
        if not CellsCoordinates._isValidExcelCell(select.start_cell):
            raise Exception(f"Start cell not valid: {select.start_cell=}")
        selection.start_cell = _CoordinateParam(select.start_cell)

        # Set end_cell
        if select.end_cell == None:
            selection.end_cell = selection.start_cell
        elif CellsCoordinates._isValidExcelCell(select.end_cell):
            selection.end_cell = _CoordinateParam(select.end_cell)
        elif CellsCoordinates._isValidExcelColumn(select.end_cell):
            select.end_cell = CellsCoordinates._get_last_NonEmpty_row(selection, select.end_cell)
            selection.end_cell = _CoordinateParam(select.end_cell)
        elif CellsCoordinates._isValidExcelRow(select.end_cell):
            select.end_cell = CellsCoordinates._get_last_NonEmpty_col(selection, select.end_cell)
            selection.end_cell = _CoordinateParam(select.end_cell)
        else:
            raise ValueError("End cell format error")

        # Set ranges
        if not(CellsCoordinates._isValidExelRange(selection.start_cell, selection.end_cell)):
            raise Exception('start_cell must be located before end_cell')
        selection.rows_range = range(selection.start_cell.row_number, selection.end_cell.row_number + 1)
        selection.columns_range = range(selection.start_cell.col_number, selection.end_cell.col_number + 1)

        return selection

    @staticmethod
    def _get_last_NonEmpty_row(cells:CellsSelection, col_letter:str) -> str:
        row = cells.start_cell.row_number
        col_number = opxl_cell.column_index_from_string(col_letter)
        while cells.worksheet.cell(row=row, column=col_number).value is not None:
            row += 1
        max_row = row - 1
        return f"{col_letter}{max_row}"

    @staticmethod
    def _get_last_NonEmpty_col(cells:CellsSelection, r:str) -> str:
        col = int(cells.start_cell.col_number)
        while cells.worksheet.cell(row=int(r), column=col).value is not None:
            col += 1
        max_col_letter = opxl_utils.get_column_letter(col - 1)
        return f"{max_col_letter}{r}"

    @staticmethod
    def _isValidExcelWorksheet(ws: OpxlWorksheet) -> bool:
        return type(ws) == OpxlWorksheet

    @staticmethod
    def _isValidExcelCell(cell: str) -> bool:
        if not type(cell) == str:
            raise TypeError("Type error")

        m = re.fullmatch(r"^([A-Z]{1,3})([1-9]\d*)$", cell)
        if not m:
            return False

        letters = m.group(1)
        digits = m.group(2)
        if not CellsCoordinates._isValidExcelColumn(letters):
            log.error("Cell's column value out of range")
            return False
        if not CellsCoordinates._isValidExcelRow(digits):
            log.error("Cell's row value out of range")
            return False

        return True

    @staticmethod
    def _isValidExcelColumn(col: str) -> bool:
        return not (
            not type(col) == str
            or not col.isalpha()
            or len(col) > len(MAX_COLUMN_EXL)
            or (len(col) == len(MAX_COLUMN_EXL) and col > MAX_COLUMN_EXL)
        )

    @staticmethod
    def _isValidExcelRow(row: str) -> bool:
        return not (not type(row) == str or not row.isnumeric() or int(row) > MAX_ROW_EXL)

    @staticmethod
    def _isValidExelRange(start:_CoordinateParam, end:_CoordinateParam) -> bool:
        return not (start.col_number > end.col_number or start.row_number > end.row_number)