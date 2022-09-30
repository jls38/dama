import logging
from dataclasses import dataclass
from typing import Optional,Tuple

from openpyxl import utils as opxl_utils, styles as opxl_styles

from DamaLib.common.excel.sheetformat.Coordinates import CellsCoordinates, DMSelect
from DamaLib.common.utils.color import ColorConverter, Color
from DamaLib.common.decorators.check import check_method_input, check_dataclass_input

log = logging.getLogger(__name__)

@check_dataclass_input
@dataclass
class FontParameters:
    """
    underline must be one of {'none','single', 'double'}
    """

    name: Optional[str] = None
    size: Optional[int | float] = None
    bold: Optional[bool] = None
    italic: Optional[bool] = None
    underline: Optional[str] = None #"none"
    strike: Optional[bool] = None
    color: Optional[str | Color] = None

    def __post_init__(self):
        if not self.color == None:
            self.color = ColorConverter(self.color).hexaCode()

        underline_param = ("none", "single", "double")
        if not (self.underline in underline_param or self.underline == None):
            raise ValueError(f"underline must be one of {underline_param=}")

@check_dataclass_input
@dataclass
class BordersParameters:
    """
    sides:
        - left | right | top | bottom | diagonal
        - vertical | horizontal
        - outline (for no color)
        - rangeOutline (for range Outline)

    border_style must be one of ['slantDashDot', 'thin', 'double', 'hair', 'dashDot', 'dashDotDot', 'mediumDashDot', 'mediumDashDotDot', 'dotted', 'mediumDashed', 'medium', 'dashed', 'thick']

    color: str (name or hexacode(#))
    """
    sides: Optional[Tuple] = None
    border_style: str = None
    color: Optional[str|Color] = None

    def __post_init__(self):
        if not self.sides == "rangeOutline":
            sides_list = (
                "left",
                "right",
                "top",
                "bottom",
                "diagonal",
                "vertical",
                "horizontal",
                "outline",
            )

            for side in self.sides:
                if not side in sides_list:
                    raise ValueError(f"Sides must be on the list {sides_list}")

        border_list = (
            "none",
            "slantDashDot",
            "thin",
            "double",
            "hair",
            "dashDot",
            "dashDotDot",
            "mediumDashDot",
            "mediumDashDotDot",
            "dotted",
            "mediumDashed",
            "medium",
            "dashed",
            "thick",
        )
        if not self.border_style in border_list:
            raise ValueError(f"Sides must be on the {border_list=}")

        if self.color is not None:
            self.color = ColorConverter(self.color).hexaCode()

@check_dataclass_input
@dataclass
class FillParameters:
    fgColor: Optional[str | Color] = None
    fill_type: Optional[str] = "solid"
    patternType: Optional[str] = None

    def __post_init__(self):
        if self.fgColor is not None:
            self.fgColor = ColorConverter(self.fgColor).hexaCode()

@check_dataclass_input
@dataclass
class AlignementParameters:
    """
    horizontal must be one of : 'centerContinuous', 'distributed', 'right', 'fill', 'left', 'general', 'justify', 'center'
    vertical must be one of : 'bottom', 'justify', 'distributed', 'center', 'top'
    """
    horizontal: Optional[str] = None
    vertical: Optional[str] = None
    wrap_text: Optional[bool] = None
    shrink_to_fit: Optional[bool] = None
    text_rotation: int = 0
    indent: int = 0

    def __post_init__(self):
        h_alignment_list = (
            "centerContinuous",
            "distributed",
            "right",
            "fill",
            "left",
            "general",
            "justify",
            "center",
        )

        if self.horizontal is not None and not self.horizontal in h_alignment_list:
            raise ValueError(f"horizontal must be None or one of {h_alignment_list=}")

        v_alignment_list = (
            "bottom",
            "justify",
            "distributed",
            "center",
            "top",
        )

        if self.vertical is not None and not self.vertical in v_alignment_list:
            raise ValueError(f"horizontal must be None or one of {v_alignment_list=}")


class Cell(object):
    def __init__(
        self, cells_selected:DMSelect
    ) -> None:
        self.selected_datas = CellsCoordinates.set_up_selection(cells_selected)
        self.worksheet = self.selected_datas.worksheet
        self.cell = self.selected_datas.worksheet.cell
        self.rows_range = self.selected_datas.rows_range
        self.cols_range = self.selected_datas.columns_range

        self._fontparam = FontParameters
        self._fillparam = FillParameters
        self._alignparam = AlignementParameters
        self._numformatparam = 'general'
        self._borderparam = BordersParameters
        self._rows_width_param = int
        self._rows_height = int

    @property
    def fillparam(self):
        return self._fillparam

    @check_method_input('')
    @fillparam.setter
    def fillparam(self, fill_parameters: FillParameters):
        self._fillparam = fill_parameters

    def fill_cell(self) -> None:
        cells_fill = {str(attr): getattr(self.fillparam, attr) for attr in self.fillparam.__dict__}
        [setattr(self.cell(row=r, column=c), 'fill', opxl_styles.PatternFill(**cells_fill)) for r in self.rows_range for c in self.cols_range]

    @property
    def fontparam(self) -> FontParameters:
        return self._fontparam

    @check_method_input('')
    @fontparam.setter
    def fontparam(self, font_parameters: FontParameters) -> None:
        self._fontparam = font_parameters

    def apply_font(self) -> None:
        font = {str(attr): getattr(self.fontparam, attr) for attr in self.fontparam.__dict__}
        [setattr(self.cell(row=r, column=c), 'font', opxl_styles.Font(**font)) for r in self.rows_range for c in self.cols_range]       

    @property
    def numformatparam (self):
        return self._numformatparam

    @check_method_input('')
    @numformatparam.setter
    def numformatparam(self, num_format:str):
        """num format: 'general' | '0' | '#,##0.00' | '#,##0.00E+00' | '#,#0.0%' | str(...)"""
        self._numformatparam = num_format

    def apply_number_format(self) -> None:
        [setattr(self.cell(row=r, column=c), 'number_format', self.numformatparam) for r in self.rows_range for c in self.cols_range]

    @property
    def alignparam(self):
        return self._alignparam

    @check_method_input('')
    @alignparam.setter
    def alignparam(self, align: AlignementParameters):
        self._alignparam = align

    def apply_alignment(self):
        cAlignment = {str(attr): getattr(self.alignparam, attr) for attr in self.alignparam.__dict__}
        [setattr(self.cell(row=r, column=c), 'alignment', opxl_styles.Alignment(**cAlignment)) for r in self.rows_range for c in self.cols_range]

    @property
    def borderparam(self):
        return self._borderparam

    @check_method_input('')
    @borderparam.setter
    def borderparam(self, borders: BordersParameters):
        self._borderparam = borders

    def apply_borders(self) -> None:
        cells_Border = {"border_style": self.borderparam.border_style, "color": self.borderparam.color}
        side = opxl_styles.Side(**cells_Border)

        match self.borderparam.sides:
            case 'rangeOutline':
                max_row = len(self.rows_range) - 1
                max_col = len(self.cols_range) - 1

                for i, r in enumerate(self.rows_range):
                    for j, c in enumerate(self.cols_range):
                        # Initializing borders
                        cell = self.cell(row=r, column=c)
                        border = opxl_styles.Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                        )
                        # Set border on the side of the selection
                        if j == 0:
                            border.left = side
                        if j == max_col:
                            border.right = side
                        if i == 0:
                            border.top = side
                        if i == max_row:
                            border.bottom = side
                        if i == 0 or i == max_row or j == 0 or j == max_col:
                            cell.border = border
            case _:
                # Define borders' format
                borders_format = {s: side for s in self.borderparam.sides}
                [setattr(self.cell(row=r, column=c), 'border', opxl_styles.Border(**borders_format)) for r in self.rows_range for c in self.cols_range]

    @property
    def cols_width_param(self):
        return self._rows_width_param

    @check_method_input('')
    @cols_width_param.setter
    def cols_width_param(self, width: int | float):
        self._rows_width_param = width
    
    def apply_cols_width(self) -> None:
        for col in self.cols_range:
            c = opxl_utils.get_column_letter(col)
            self.worksheet.column_dimensions[c].width = self.cols_width_param

    @property
    def rows_height_param(self):
        return self._rows_height

    @check_method_input('')
    @rows_height_param.setter
    def rows_height_param(self, height: int | float):
        self._rows_height = height

    def apply_rows_height(self) -> None:
        [setattr(self.worksheet.row_dimensions[row], 'height', self.rows_height_param) for row in self.rows_range] 

    def merge_cells(self) -> None:
        # Check if less than 2 cells are not empty
        cells_not_empty = 0
        for r in self.rows_range:
            for c in self.cols_range:
                if not self.cell(row=r, column=c).value == None:
                    cells_not_empty += 1
                if cells_not_empty >= 2:
                    log.error("Impossile to merge cells")
                    break
            else:
                continue
            break
        else:
            self.worksheet.merge_cells(f"{self.selected_datas.start_cell.coordinate}:{self.selected_datas.end_cell.coordinate}")

    def freeze_cell(self) -> None:
        self.worksheet.freeze_panes = self.selected_datas.start_cell.coordinate