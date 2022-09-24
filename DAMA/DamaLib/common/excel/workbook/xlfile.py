import logging
from os import path as os_path

from pandas import DataFrame as pd_DataFrame, ExcelWriter as pd_ExcelWriter
from openpyxl import load_workbook as opxl_loadWB
from openpyxl.worksheet.worksheet import Worksheet as opxl_Worksheet
from openpyxl.workbook.workbook import Workbook as opxl_Workbook

log = logging.getLogger(__name__)

class xlfile (object):

    def __init__() -> None:
        pass

    def active_workbook() -> opxl_Workbook:
        wb = opxl_Workbook.active
        return wb

    def load_workbook(path:str) -> opxl_Workbook:
        wb = opxl_loadWB(path)
        return wb

    def active_worksheet() -> opxl_Worksheet:
        wb = opxl_Workbook()
        ws = wb.active
        return ws

    def load_worksheet(wb:opxl_Workbook, sheet:str) -> opxl_Worksheet :
        return wb[sheet]

    def add_to_xlsx (filepath:str, sheet:str, df:pd_DataFrame, start_row:int, start_col:int, Header:bool= True, Index:bool= False) -> None:
        if os_path.isfile(filepath) == False:
            raise TypeError('No file specified')

        with pd_ExcelWriter(
            filepath, mode='a', 
            engine='openpyxl', 
            if_sheet_exists='overlay'
            ) as writer:
            df.to_excel(
                writer, 
                sheet_name = sheet, 
                header = Header,
                index = Index, 
                startrow = start_row, 
                startcol = start_col
                )

        log.info('Dataframe add to : %s', filepath)

    def get_xls_sheetsList (filepath:str) -> list[str]:
        workbook = opxl_loadWB(filepath)
        return workbook.sheetnames

    def remove_sheet(filepath:str, sheet:str) -> None:
        workbook = opxl_loadWB(filepath)
        std = workbook.get_sheet_by_name(sheet)
        workbook.remove_sheet(std)
        workbook.save(filepath)
        log.info('%s removed from %s', sheet, filepath)

    def remove_charts(filepath:str, sheet:str, chart_number:int|None) -> None:
        wb = xlfile.load_workbook(filepath)
        ws = xlfile.load_worksheet(wb, sheet)

        if chart_number == None:
            for s in range(0, len(ws._charts)):
                del ws._charts[s]
        else:
            del ws._charts[chart_number]
        wb.save(filepath)
